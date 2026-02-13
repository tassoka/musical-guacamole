import json
import os
import re
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

from flask import (
    Flask,
    jsonify,
    render_template,
    request,
    send_file,
)
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import xlrd

IMPORT_EXPECTED_HEADERS = [
    "категория",
    "наименование",
    "опт",
    "цена продажи",
    "остаток",
    "gtin",
    "дата поступления",
]

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "inventory.db")

app = Flask(__name__)
app.config["SECRET_KEY"] = "change-me"

SUPPLIER_AUTO_REFRESH_MINUTES = 30


def get_db_connection():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def _resolve_google_sheet_url(url: str) -> Optional[str]:
    parsed = urlparse(url)
    if "docs.google.com" not in parsed.netloc:
        return None
    path_match = re.search(r"/d/([a-zA-Z0-9-_]+)", parsed.path)
    if not path_match:
        return None
    document_id = path_match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{document_id}/export?format=xlsx"


def _resolve_supplier_source_url(url: str) -> str:
    url = url.strip()
    if not url:
        raise ValueError("Ссылка обязательна.")

    parsed = urlparse(url)
    if not parsed.scheme:
        raise ValueError("Укажите корректный URL со схемой (http/https).")

    google_url = _resolve_google_sheet_url(url)
    if google_url:
        return google_url

    return url


def _load_xls_workbook(data: bytes) -> Workbook:
    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as error:
        raise ValueError(
            "Не удалось прочитать прайс-лист в формате .xls."
        ) from error

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet in book.sheets():
        title = (sheet.name or "Лист").strip() or "Лист"
        worksheet = workbook.create_sheet(title=title[:31])
        for row_index in range(sheet.nrows):
            row_values = []
            for col_index in range(sheet.ncols):
                cell_value = sheet.cell_value(row_index, col_index)
                cell_type = sheet.cell_type(row_index, col_index)
                if cell_type == xlrd.XL_CELL_DATE:
                    try:
                        cell_value = xlrd.xldate_as_datetime(cell_value, book.datemode)
                    except Exception:
                        pass
                elif cell_type == xlrd.XL_CELL_NUMBER:
                    if float(cell_value).is_integer():
                        cell_value = int(cell_value)
                elif cell_type in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    cell_value = None
                row_values.append(cell_value)
            worksheet.append(row_values if row_values else [None])

    if not workbook.sheetnames:
        workbook.create_sheet("Лист1")

    return workbook


def _load_workbook_from_bytes(data: bytes, file_name: Optional[str] = None) -> Workbook:
    if not data:
        raise ValueError("Файл прайс-листа пустой или не загружен.")

    if file_name and file_name.lower().endswith(".xls"):
        return _load_xls_workbook(data)

    stream = BytesIO(data)
    try:
        return load_workbook(stream, data_only=True)
    except Exception as error:
        raise ValueError(
            "Не удалось прочитать прайс-лист. Поддерживаются файлы Excel (например, .xlsx или .xls)."
        ) from error


def _download_supplier_workbook(url: str) -> Workbook:
    resolved = _resolve_supplier_source_url(url)
    try:
        response = requests.get(resolved, timeout=30)
    except requests.RequestException as error:
        raise ValueError(f"Не удалось загрузить файл: {error}") from error

    if response.status_code >= 400:
        raise ValueError("Сервер вернул ошибку при загрузке файла.")

    return _load_workbook_from_bytes(response.content)


def _format_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def _normalize_text(value: Optional[Any]) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_multiplier(value: Optional[Any], default: float = 1.0) -> float:
    if value is None:
        return default
    text = str(value).strip().replace(" ", "")
    if not text:
        return default
    text = text.replace(",", ".")
    try:
        multiplier = float(text)
    except ValueError as error:
        raise ValueError("Некорректное значение курса валюты.") from error
    if multiplier <= 0:
        raise ValueError("Курс валюты должен быть больше нуля.")
    return multiplier


def _sheet_preview(workbook, limit: int = 100):
    sheets = []
    for worksheet in workbook.worksheets:
        columns = []
        rows: List[List[str]] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            formatted = [_format_cell(cell) for cell in row]
            if row_index == 1:
                for idx, header in enumerate(formatted, start=1):
                    columns.append({
                        "letter": get_column_letter(idx),
                        "header": header,
                    })
            elif len(formatted) > len(columns):
                for idx in range(len(columns) + 1, len(formatted) + 1):
                    columns.append(
                        {
                            "letter": get_column_letter(idx),
                            "header": "",
                        }
                    )
            if row_index > limit:
                break
            rows.append(formatted)
        sheets.append({
            "name": worksheet.title,
            "columns": columns,
            "rows": rows,
        })
    return sheets


def _apply_price_multiplier(value: Optional[str], multiplier: float) -> str:
    text = _normalize_text(value)
    if not text:
        return text
    normalized = text.replace(" ", "").replace(",", ".")
    try:
        numeric = float(normalized)
    except ValueError:
        return text
    result = numeric * multiplier
    if result.is_integer():
        return str(int(result))
    return f"{result:.2f}".rstrip("0").rstrip(".")


def _get_handler_sheets(connection: sqlite3.Connection, handler_id: int) -> List[Dict[str, Any]]:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT
            id,
            sheet_name,
            position,
            column_name,
            column_wholesale,
            column_recommended,
            column_stock,
            enabled
        FROM supplier_handler_sheets
        WHERE handler_id = ?
        ORDER BY position ASC, id ASC
        """,
        (handler_id,),
    )
    sheets = []
    for row in cursor.fetchall():
        sheets.append(
            {
                "id": row["id"],
                "sheet_name": row["sheet_name"],
                "position": row["position"],
                "column_name": row["column_name"],
                "column_wholesale": row["column_wholesale"],
                "column_recommended": row["column_recommended"],
                "column_stock": row["column_stock"],
                "enabled": bool(row["enabled"]),
            }
        )
    return sheets


def _get_supplier_handler(handler_id: int):
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT
            id,
            title,
            source_url,
            source_file_name,
            source_file,
            sheet_name,
            column_name,
            column_wholesale,
            column_recommended,
            column_stock,
            wholesale_multiplier,
            recommended_multiplier,
            created_at,
            last_refreshed_at
        FROM supplier_handlers
        WHERE id = ?
        """,
        (handler_id,),
    )
    row = cursor.fetchone()
    if row is None:
        connection.close()
        return None
    record = dict(row)
    record["sheets"] = _get_handler_sheets(connection, handler_id)
    connection.close()
    blob = record.get("source_file")
    if isinstance(blob, memoryview):
        record["source_file"] = blob.tobytes()
    return record


def _serialize_handler(handler: Dict[str, Any]) -> Dict[str, Any]:
    serialized = dict(handler)
    serialized["has_file"] = bool(serialized.get("source_file"))
    serialized.pop("source_file", None)
    return serialized


def _load_handler_workbook(handler) -> Workbook:
    source_file = handler.get("source_file")
    if source_file:
        file_name = handler.get("source_file_name")
        if isinstance(source_file, memoryview):
            source_file = source_file.tobytes()
        return _load_workbook_from_bytes(source_file, file_name=file_name)
    source_url = (handler.get("source_url") or "").strip()
    if not source_url:
        raise ValueError("Не указан источник прайс-листа.")
    return _download_supplier_workbook(source_url)


def _handler_has_mapping(handler: Dict[str, str]) -> bool:
    sheets = handler.get("sheets")
    if isinstance(sheets, list) and sheets:
        for sheet in sheets:
            if not sheet:
                continue
            if not sheet.get("enabled", True):
                continue
            if (sheet.get("sheet_name") or "").strip() and (sheet.get("column_name") or "").strip():
                return True
        return False
    required = ["sheet_name", "column_name"]
    return all((handler.get(field) or "").strip() for field in required)


def _safe_column_letter(value: Optional[str]) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def _validate_column_letter(value: Optional[str], label: str, required: bool) -> Optional[str]:
    column = _safe_column_letter(value)
    if not column:
        if required:
            raise ValueError(label)
        return None
    if not re.fullmatch(r"[A-Z]+", column):
        raise ValueError(label)
    try:
        column_index_from_string(column)
    except ValueError as error:
        raise ValueError(label) from error
    return column


def _parse_sheet_configs(raw_value: Any, *, require_selection: bool = True) -> List[Dict[str, Any]]:
    if isinstance(raw_value, str):
        raw_text = raw_value.strip()
        if not raw_text:
            data = []
        else:
            try:
                data = json.loads(raw_text)
            except json.JSONDecodeError as error:
                raise ValueError("Не удалось разобрать выбранные листы.") from error
    elif isinstance(raw_value, list):
        data = raw_value
    else:
        data = []

    configs: List[Dict[str, Any]] = []
    seen_names = set()
    for index, entry in enumerate(data):
        if not isinstance(entry, dict):
            continue
        enabled = entry.get("enabled", True)
        if not enabled:
            continue
        sheet_name = (entry.get("sheet_name") or entry.get("name") or "").strip()
        if not sheet_name:
            continue
        normalized_key = sheet_name.lower()
        if normalized_key in seen_names:
            continue
        mapping_source = entry.get("mapping") or {}
        if not isinstance(mapping_source, dict):
            mapping_source = {}
        try:
            column_name = _validate_column_letter(
                mapping_source.get("name"),
                f"Укажите колонку с наименованием для листа «{sheet_name}».",
                required=True,
            )
            column_wholesale = _validate_column_letter(
                mapping_source.get("wholesale_price"),
                f"Некорректная колонка «Опт» для листа «{sheet_name}».",
                required=False,
            )
            column_recommended = _validate_column_letter(
                mapping_source.get("recommended_price"),
                f"Некорректная колонка «РРЦ» для листа «{sheet_name}».",
                required=False,
            )
            column_stock = _validate_column_letter(
                mapping_source.get("stock"),
                f"Некорректная колонка «Остаток» для листа «{sheet_name}».",
                required=False,
            )
        except ValueError as error:
            raise ValueError(str(error)) from error

        configs.append(
            {
                "sheet_name": sheet_name,
                "position": entry.get("position", index),
                "column_name": column_name,
                "column_wholesale": column_wholesale,
                "column_recommended": column_recommended,
                "column_stock": column_stock,
                "enabled": True,
            }
        )
        seen_names.add(normalized_key)

    configs.sort(key=lambda item: (item.get("position", 0), item["sheet_name"].lower()))

    if require_selection and not configs:
        raise ValueError("Выберите листы и настройте сопоставление колонок.")

    return configs


def _get_cached_supplier_rows(connection, handler_id: int) -> List[Dict[str, Any]]:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT
            sheet_name,
            row_kind,
            position,
            category,
            name,
            wholesale_price,
            recommended_price,
            stock
        FROM supplier_items
        WHERE handler_id = ?
        ORDER BY position ASC, id ASC
        """,
        (handler_id,),
    )
    rows = [
        {
            "sheet": row["sheet_name"],
            "kind": row["row_kind"] or "item",
            "position": row["position"],
            "category": row["category"],
            "name": row["name"],
            "wholesale_price": row["wholesale_price"],
            "recommended_price": row["recommended_price"],
            "stock": row["stock"],
        }
        for row in cursor.fetchall()
    ]
    return rows


def _get_recent_supplier_changes(
    connection: sqlite3.Connection, handler_id: int, limit: int = 50
) -> List[Dict[str, Any]]:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT
            id,
            sheet_name,
            item_name,
            category,
            field,
            old_value,
            new_value,
            changed_at
        FROM supplier_item_changes
        WHERE handler_id = ?
        ORDER BY datetime(changed_at) DESC, id DESC
        LIMIT ?
        """,
        (handler_id, limit),
    )
    return [dict(row) for row in cursor.fetchall()]


def _get_all_recent_changes(connection: sqlite3.Connection, limit: int = 100) -> List[Dict[str, Any]]:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT
            c.id,
            c.handler_id,
            h.title as handler_title,
            c.sheet_name,
            c.item_name,
            c.category,
            c.field,
            c.old_value,
            c.new_value,
            c.changed_at
        FROM supplier_item_changes c
        JOIN supplier_handlers h ON c.handler_id = h.id
        ORDER BY datetime(c.changed_at) DESC, c.id DESC
        LIMIT ?
        """,
        (limit,),
    )
    return [dict(row) for row in cursor.fetchall()]


def _refresh_handler_rows(handler) -> List[Dict[str, Any]]:
    if not _handler_has_mapping(handler):
        raise ValueError("Не выбраны колонки для прайс-листа.")

    workbook = _load_handler_workbook(handler)
    try:
        wholesale_multiplier = float(handler.get("wholesale_multiplier") or 1.0)
    except (TypeError, ValueError):
        wholesale_multiplier = 1.0
    try:
        recommended_multiplier = float(handler.get("recommended_multiplier") or 1.0)
    except (TypeError, ValueError):
        recommended_multiplier = 1.0
    mapping: Dict[str, Dict[str, Any]] = {}
    sheets = handler.get("sheets")
    if isinstance(sheets, list) and sheets:
        for sheet in sheets:
            if not sheet or not sheet.get("enabled", True):
                continue
            sheet_name = (sheet.get("sheet_name") or "").strip()
            column_name = _safe_column_letter(sheet.get("column_name"))
            if not sheet_name or not column_name:
                continue
            mapping[sheet_name] = {
                "enabled": True,
                "mapping": {
                    "name": column_name,
                    "wholesale_price": _safe_column_letter(sheet.get("column_wholesale")),
                    "recommended_price": _safe_column_letter(sheet.get("column_recommended")),
                    "stock": _safe_column_letter(sheet.get("column_stock")),
                },
            }
    else:
        sheet_name = (handler.get("sheet_name") or "").strip()
        column_name = _safe_column_letter(handler.get("column_name"))
        if sheet_name and column_name:
            mapping[sheet_name] = {
                "enabled": True,
                "mapping": {
                    "name": column_name,
                    "wholesale_price": _safe_column_letter(handler.get("column_wholesale")),
                    "recommended_price": _safe_column_letter(handler.get("column_recommended")),
                    "stock": _safe_column_letter(handler.get("column_stock")),
                },
            }

    if not mapping:
        raise ValueError("Не выбраны листы или колонки для прайс-листа.")

    rows = _extract_mapped_rows(workbook, mapping, limit=500)

    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT
            sheet_name,
            row_kind,
            category,
            name,
            wholesale_price,
            recommended_price,
            stock
        FROM supplier_items
        WHERE handler_id = ?
        """,
        (handler["id"],),
    )
    existing_items: Dict[Any, Dict[str, str]] = {}

    def _make_key(sheet: Any, category: Any, name: Any) -> Any:
        return (
            _normalize_text(sheet).lower(),
            _normalize_text(category).lower(),
            _normalize_text(name).lower(),
        )

    for row in cursor.fetchall():
        kind = (row["row_kind"] or "item").strip()
        if kind != "item":
            continue
        name_value = _normalize_text(row["name"])
        if not name_value:
            continue
        key = _make_key(row["sheet_name"], row["category"], name_value)
        if key in existing_items:
            continue
        existing_items[key] = {
            "wholesale_price": _normalize_text(row["wholesale_price"]),
            "recommended_price": _normalize_text(row["recommended_price"]),
            "stock": _normalize_text(row["stock"]),
        }

    cursor.execute("DELETE FROM supplier_items WHERE handler_id = ?", (handler["id"],))
    timestamp = datetime.utcnow().isoformat(timespec="seconds")
    rows_with_positions: List[Dict[str, Any]] = []
    change_entries = []
    for position, row in enumerate(rows):
        normalized = dict(row)
        normalized.setdefault("kind", "item")
        normalized["position"] = position
        if normalized.get("kind") == "item":
            normalized["wholesale_price"] = _apply_price_multiplier(
                normalized.get("wholesale_price"), wholesale_multiplier
            )
            normalized["recommended_price"] = _apply_price_multiplier(
                normalized.get("recommended_price"), recommended_multiplier
            )
        rows_with_positions.append(normalized)
        cursor.execute(
            """
            INSERT INTO supplier_items (
                handler_id,
                sheet_name,
                row_kind,
                position,
                category,
                name,
                wholesale_price,
                recommended_price,
                stock,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                handler["id"],
                normalized.get("sheet"),
                normalized.get("kind"),
                normalized.get("position"),
                normalized.get("category"),
                normalized.get("name"),
                normalized.get("wholesale_price"),
                normalized.get("recommended_price"),
                normalized.get("stock"),
                timestamp,
            ),
        )
        if normalized.get("kind") == "item":
            name_value = _normalize_text(normalized.get("name"))
            if name_value:
                key = _make_key(
                    normalized.get("sheet"),
                    normalized.get("category"),
                    name_value,
                )
                previous = existing_items.get(key)
                if previous:
                    for field in ("wholesale_price", "recommended_price", "stock"):
                        new_value = _normalize_text(normalized.get(field))
                        old_value = previous.get(field, "")
                        if new_value != old_value:
                            change_entries.append(
                                (
                                    handler["id"],
                                    normalized.get("sheet"),
                                    normalized.get("name"),
                                    normalized.get("category"),
                                    field,
                                    old_value,
                                    new_value,
                                    timestamp,
                                )
                            )
    cursor.execute(
        "UPDATE supplier_handlers SET last_refreshed_at = ? WHERE id = ?",
        (timestamp, handler["id"]),
    )
    if change_entries:
        cursor.executemany(
            """
            INSERT INTO supplier_item_changes (
                handler_id,
                sheet_name,
                item_name,
                category,
                field,
                old_value,
                new_value,
                changed_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            change_entries,
        )
    cursor.execute(
        """
        DELETE FROM supplier_item_changes
        WHERE handler_id = ?
          AND id NOT IN (
              SELECT id
              FROM supplier_item_changes
              WHERE handler_id = ?
              ORDER BY datetime(changed_at) DESC, id DESC
              LIMIT 200
          )
        """,
        (handler["id"], handler["id"]),
    )
    connection.commit()
    connection.close()

    return rows_with_positions


def _extract_mapped_rows(workbook: Workbook, selections: Dict[str, Dict[str, str]], limit: int = 200):
    result = []
    field_definitions = {
        "name": {"label": "Наименование", "required": True},
        "wholesale_price": {"label": "Опт", "required": False},
        "recommended_price": {"label": "РРЦ", "required": False},
        "stock": {"label": "Остаток", "required": False},
    }

    for sheet_name, config in selections.items():
        if not config.get("enabled"):
            continue
        mapping = config.get("mapping") or {}
        if not mapping:
            continue
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        if worksheet is None:
            continue
        indexes: Dict[str, Optional[int]] = {}
        valid = True
        for key, definition in field_definitions.items():
            column_letter = (mapping.get(key) or "").strip().upper()
            if not column_letter:
                if definition["required"]:
                    valid = False
                    break
                indexes[key] = None
                continue
            try:
                indexes[key] = column_index_from_string(column_letter)
            except ValueError:
                valid = False
                break
        if not valid or indexes.get("name") is None:
            continue

        any_optional_mapped = any(
            indexes.get(field) is not None for field in field_definitions if field != "name"
        )

        current_category: Optional[str] = None
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue
            record: Dict[str, Any] = {"sheet": sheet_name}
            for key, column_index in indexes.items():
                if column_index is None:
                    record[key] = None
                else:
                    value = row[column_index - 1] if column_index - 1 < len(row) else None
                    record[key] = _format_cell(value)

            name_value = (record.get("name") or "").strip()
            other_fields = [
                record.get(field)
                for field in field_definitions
                if field != "name" and indexes.get(field) is not None
            ]
            has_other_values = any(value for value in other_fields)

            if not name_value and not has_other_values:
                continue

            if name_value and not has_other_values and any_optional_mapped:
                current_category = name_value
                result.append(
                    {
                        "sheet": sheet_name,
                        "kind": "category",
                        "category": current_category,
                    }
                )
            else:
                record["kind"] = "item"
                record["category"] = current_category
                result.append(record)

            if len(result) >= limit:
                return result
    return result


def init_db():
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT,
            name TEXT NOT NULL,
            wholesale_price REAL NOT NULL DEFAULT 0,
            sale_price REAL,
            quantity INTEGER NOT NULL DEFAULT 0,
            gtin TEXT,
            date_received TEXT NOT NULL
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS inventory_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            change_amount INTEGER NOT NULL,
            action TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(item_id) REFERENCES items(id) ON DELETE CASCADE
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS supplier_handlers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            source_url TEXT NOT NULL,
            source_file_name TEXT,
            source_file BLOB,
            sheet_name TEXT,
            column_name TEXT,
            column_wholesale TEXT,
            column_recommended TEXT,
            column_stock TEXT,
            wholesale_multiplier REAL NOT NULL DEFAULT 1,
            recommended_multiplier REAL NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            last_refreshed_at TEXT
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS supplier_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            handler_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            row_kind TEXT NOT NULL DEFAULT 'item',
            position INTEGER NOT NULL DEFAULT 0,
            category TEXT,
            name TEXT,
            wholesale_price TEXT,
            recommended_price TEXT,
            stock TEXT,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(handler_id) REFERENCES supplier_handlers(id) ON DELETE CASCADE
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS supplier_item_changes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            handler_id INTEGER NOT NULL,
            sheet_name TEXT,
            item_name TEXT,
            category TEXT,
            field TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            changed_at TEXT NOT NULL,
            FOREIGN KEY(handler_id) REFERENCES supplier_handlers(id) ON DELETE CASCADE
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS supplier_handler_sheets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            handler_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            position INTEGER NOT NULL DEFAULT 0,
            column_name TEXT NOT NULL,
            column_wholesale TEXT,
            column_recommended TEXT,
            column_stock TEXT,
            enabled INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY(handler_id) REFERENCES supplier_handlers(id) ON DELETE CASCADE
        )
        """
    )
    cursor.execute("PRAGMA table_info(supplier_handlers)")
    existing_columns = {row[1] for row in cursor.fetchall()}
    alterations = [
        ("sheet_name", "TEXT"),
        ("column_name", "TEXT"),
        ("column_wholesale", "TEXT"),
        ("column_recommended", "TEXT"),
        ("column_stock", "TEXT"),
        ("source_file_name", "TEXT"),
        ("source_file", "BLOB"),
        ("wholesale_multiplier", "REAL DEFAULT 1"),
        ("recommended_multiplier", "REAL DEFAULT 1"),
        ("last_refreshed_at", "TEXT"),
    ]
    for column, column_type in alterations:
        if column not in existing_columns:
            cursor.execute(
                f"ALTER TABLE supplier_handlers ADD COLUMN {column} {column_type}"
            )
    cursor.execute("PRAGMA table_info(supplier_items)")
    existing_supplier_item_columns = {row[1] for row in cursor.fetchall()}
    supplier_item_alterations = [
        ("row_kind", "TEXT DEFAULT 'item'"),
        ("position", "INTEGER DEFAULT 0"),
        ("category", "TEXT"),
    ]
    for column, column_type in supplier_item_alterations:
        if column not in existing_supplier_item_columns:
            cursor.execute(
                f"ALTER TABLE supplier_items ADD COLUMN {column} {column_type}"
            )

    cursor.execute(
        """
        SELECT id, sheet_name, column_name, column_wholesale, column_recommended, column_stock
        FROM supplier_handlers
        """
    )
    handlers_for_migration = cursor.fetchall()
    for handler in handlers_for_migration:
        cursor.execute(
            "SELECT 1 FROM supplier_handler_sheets WHERE handler_id = ? LIMIT 1",
            (handler["id"],),
        )
        if cursor.fetchone():
            continue
        sheet_name = (handler["sheet_name"] or "").strip()
        column_name = (handler["column_name"] or "").strip()
        if not sheet_name or not column_name:
            continue
        cursor.execute(
            """
            INSERT INTO supplier_handler_sheets (
                handler_id,
                sheet_name,
                position,
                column_name,
                column_wholesale,
                column_recommended,
                column_stock,
                enabled
            )
            VALUES (?, ?, 0, ?, ?, ?, ?, 1)
            """,
            (
                handler["id"],
                sheet_name,
                column_name,
                handler["column_wholesale"],
                handler["column_recommended"],
                handler["column_stock"],
            ),
        )
    connection.commit()
    connection.close()


def serialize_item(row):
    return {
        "id": row["id"],
        "category": row["category"],
        "name": row["name"],
        "wholesale_price": row["wholesale_price"],
        "sale_price": row["sale_price"],
        "quantity": row["quantity"],
        "gtin": row["gtin"],
        "date_received": row["date_received"],
    }


def log_history(item_id: int, change: int, action: str, connection: Optional[sqlite3.Connection] = None) -> None:
    close_after = connection is None
    if connection is None:
        connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        "INSERT INTO inventory_history (item_id, change_amount, action, created_at) VALUES (?, ?, ?, ?)",
        (item_id, change, action, datetime.utcnow().isoformat(timespec="seconds")),
    )
    if close_after:
        connection.commit()
        connection.close()


@app.route("/")
def index():
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM items ORDER BY name")
    items = [serialize_item(row) for row in cursor.fetchall()]
    connection.close()
    return render_template("index.html", items=items)


@app.route("/suppliers")
@app.route("/suppliers/<int:handler_id>")
def suppliers(handler_id: Optional[int] = None):
    return render_template("suppliers.html", active_handler_id=handler_id)


@app.route("/changes")
def changes_log():
    connection = get_db_connection()
    changes = _get_all_recent_changes(connection, limit=200)
    connection.close()
    return render_template("changes.html", changes=changes)


@app.route("/items", methods=["POST"])
def create_item():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Наименование обязательно."}), 400

    date_received = data.get("date_received") or datetime.utcnow().date().isoformat()
    try:
        datetime.fromisoformat(date_received)
    except ValueError:
        return jsonify({"error": "Неверный формат даты."}), 400

    wholesale_price = data.get("wholesale_price")
    if wholesale_price in (None, ""):
        return jsonify({"error": "Поле 'опт' обязательно."}), 400

    try:
        wholesale_price = float(wholesale_price)
    except ValueError:
        return jsonify({"error": "Поле 'опт' должно быть числом."}), 400

    sale_price = data.get("sale_price")
    if sale_price in (None, ""):
        sale_price = None
    else:
        try:
            sale_price = float(sale_price)
        except ValueError:
            return jsonify({"error": "Цена продажи должна быть числом."}), 400

    quantity = data.get("quantity", 0)
    if quantity in (None, ""):
        quantity = 0
    try:
        quantity = int(quantity)
    except (TypeError, ValueError):
        return jsonify({"error": "Количество должно быть целым числом."}), 400

    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        """
        INSERT INTO items (category, name, wholesale_price, sale_price, quantity, gtin, date_received)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            data.get("category"),
            name,
            wholesale_price,
            sale_price,
            quantity,
            data.get("gtin"),
            date_received,
        ),
    )
    item_id = cursor.lastrowid
    connection.commit()
    connection.close()

    if quantity:
        log_history(item_id, quantity, "Первоначальный остаток")

    return jsonify({"item": get_item_by_id(item_id)})


def get_item_by_id(item_id: int):
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM items WHERE id = ?", (item_id,))
    row = cursor.fetchone()
    connection.close()
    if row is None:
        return None
    return serialize_item(row)


@app.route("/items/<int:item_id>", methods=["PUT"])
def update_item(item_id: int):
    data = request.get_json(force=True)
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM items WHERE id = ?", (item_id,))
    row = cursor.fetchone()
    if row is None:
        connection.close()
        return jsonify({"error": "Товар не найден."}), 404

    updated = {
        "category": data.get("category"),
        "name": (data.get("name") or "").strip(),
        "wholesale_price": data.get("wholesale_price"),
        "sale_price": data.get("sale_price"),
        "quantity": data.get("quantity"),
        "gtin": data.get("gtin"),
        "date_received": data.get("date_received"),
    }

    if not updated["name"]:
        connection.close()
        return jsonify({"error": "Наименование обязательно."}), 400

    try:
        updated["wholesale_price"] = float(updated["wholesale_price"])
    except (TypeError, ValueError):
        connection.close()
        return jsonify({"error": "Поле 'опт' должно быть числом."}), 400

    sale_price_value = updated["sale_price"]
    if sale_price_value in (None, ""):
        updated["sale_price"] = None
    else:
        try:
            updated["sale_price"] = float(sale_price_value)
        except (TypeError, ValueError):
            connection.close()
            return jsonify({"error": "Цена продажи должна быть числом."}), 400

    if updated["quantity"] in (None, ""):
        updated["quantity"] = 0
    try:
        updated["quantity"] = int(updated["quantity"])
    except (TypeError, ValueError):
        connection.close()
        return jsonify({"error": "Количество должно быть целым числом."}), 400

    date_received = updated["date_received"] or datetime.utcnow().date().isoformat()
    try:
        datetime.fromisoformat(date_received)
    except ValueError:
        connection.close()
        return jsonify({"error": "Неверный формат даты."}), 400
    updated["date_received"] = date_received

    old_quantity = row["quantity"]
    cursor.execute(
        """
        UPDATE items
        SET category = ?, name = ?, wholesale_price = ?, sale_price = ?, quantity = ?, gtin = ?, date_received = ?
        WHERE id = ?
        """,
        (
            updated["category"],
            updated["name"],
            updated["wholesale_price"],
            updated["sale_price"],
            updated["quantity"],
            updated["gtin"],
            updated["date_received"],
            item_id,
        ),
    )
    connection.commit()
    connection.close()

    delta = updated["quantity"] - old_quantity
    if delta:
        log_history(item_id, delta, "Изменение вручную")

    return jsonify({"item": get_item_by_id(item_id)})


@app.route("/items/<int:item_id>", methods=["DELETE"])
def delete_item(item_id: int):
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT id FROM items WHERE id = ?", (item_id,))
    row = cursor.fetchone()
    if row is None:
        connection.close()
        return jsonify({"error": "Товар не найден."}), 404
    cursor.execute("DELETE FROM items WHERE id = ?", (item_id,))
    connection.commit()
    connection.close()
    return jsonify({"status": "ok"})


@app.route("/items/<int:item_id>/adjust", methods=["POST"])
def adjust_item(item_id: int):
    data = request.get_json(force=True)
    try:
        delta = int(data.get("delta"))
    except (TypeError, ValueError):
        return jsonify({"error": "Некорректное изменение количества."}), 400

    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT quantity FROM items WHERE id = ?", (item_id,))
    row = cursor.fetchone()
    if row is None:
        connection.close()
        return jsonify({"error": "Товар не найден."}), 404

    new_quantity = row["quantity"] + delta
    cursor.execute("UPDATE items SET quantity = ? WHERE id = ?", (new_quantity, item_id))
    connection.commit()
    connection.close()

    if delta:
        action = "Приход" if delta > 0 else "Расход"
        log_history(item_id, delta, action)

    return jsonify({"quantity": new_quantity})


@app.route("/items/<int:item_id>/history")
def item_history(item_id: int):
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        "SELECT change_amount, action, created_at FROM inventory_history WHERE item_id = ? ORDER BY created_at DESC",
        (item_id,),
    )
    history = [dict(row) for row in cursor.fetchall()]
    connection.close()
    return jsonify({"history": history})


@app.route("/api/suppliers", methods=["GET", "POST"])
def supplier_handlers():
    if request.method == "GET":
        summary_only = request.args.get("summary")
        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT
                id,
                title,
                source_url,
                source_file_name,
                source_file,
                sheet_name,
                column_name,
                column_wholesale,
                column_recommended,
                column_stock,
                wholesale_multiplier,
                recommended_multiplier,
                created_at,
                last_refreshed_at
            FROM supplier_handlers
            ORDER BY created_at DESC
            """
        )
        records = cursor.fetchall()

        if summary_only:
            handlers = [
                {
                    "id": row["id"],
                    "title": row["title"],
                    "created_at": row["created_at"],
                    "last_refreshed_at": row["last_refreshed_at"],
                }
                for row in records
            ]
            connection.close()
            return jsonify({"handlers": handlers})

        handlers: List[Dict[str, Any]] = []
        now = datetime.utcnow()
        for row in records:
            handler = dict(row)
            handler["sheets"] = _get_handler_sheets(connection, handler["id"])
            rows: Optional[List[Dict[str, Any]]] = None
            if _handler_has_mapping(handler):
                should_refresh = False
                refreshed_at_raw = handler.get("last_refreshed_at")
                if not refreshed_at_raw:
                    should_refresh = True
                else:
                    try:
                        refreshed_at = datetime.fromisoformat(refreshed_at_raw)
                        if refreshed_at <= now - timedelta(minutes=SUPPLIER_AUTO_REFRESH_MINUTES):
                            should_refresh = True
                    except ValueError:
                        should_refresh = True
                if should_refresh:
                    try:
                        rows = _refresh_handler_rows(handler)
                        handler = _get_supplier_handler(handler["id"]) or handler
                    except ValueError as error:
                        handler["error"] = str(error)
            if rows is None:
                rows = _get_cached_supplier_rows(connection, handler["id"])
            handler["rows"] = rows
            handler["changes"] = _get_recent_supplier_changes(connection, handler["id"])
            handlers.append(_serialize_handler(handler))
        connection.close()
        return jsonify({"handlers": handlers})

    content_type = (request.content_type or "").lower()
    use_form = "multipart/form-data" in content_type or "application/x-www-form-urlencoded" in content_type

    if use_form:
        data_source = request.form
        file_storage = request.files.get("source_file")
        file_bytes = None
        file_name = ""
        if file_storage and file_storage.filename:
            file_bytes = file_storage.read()
            file_name = file_storage.filename or ""
    else:
        data_source = request.get_json(force=True) or {}
        file_bytes = None
        file_name = ""

    title = (data_source.get("title") or "").strip()
    source_url = (data_source.get("source_url") or "").strip()
    sheet_configs_raw = data_source.get("sheets")
    try:
        sheet_configs = _parse_sheet_configs(sheet_configs_raw, require_selection=True)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    try:
        wholesale_multiplier = _parse_multiplier(
            data_source.get("wholesale_multiplier"), default=1.0
        )
    except ValueError as error:
        return jsonify({"error": f"Курс для опта: {error}"}), 400

    try:
        recommended_multiplier = _parse_multiplier(
            data_source.get("recommended_multiplier"), default=1.0
        )
    except ValueError as error:
        return jsonify({"error": f"Курс для РРЦ: {error}"}), 400

    if not title:
        return jsonify({"error": "Название обработчика обязательно."}), 400
    if not source_url and not file_bytes:
        return jsonify({"error": "Укажите ссылку или загрузите файл."}), 400

    try:
        if file_bytes:
            workbook = _load_workbook_from_bytes(file_bytes, file_name=file_name)
        else:
            workbook = _download_supplier_workbook(source_url)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    sheet_names = set(workbook.sheetnames)
    for config in sheet_configs:
        if config["sheet_name"] not in sheet_names:
            return jsonify({"error": f"Лист «{config['sheet_name']}» не найден в файле."}), 400

    primary_sheet = sheet_configs[0]

    connection = get_db_connection()
    cursor = connection.cursor()
    created_at = datetime.utcnow().isoformat(timespec="seconds")
    cursor.execute(
        """
        INSERT INTO supplier_handlers (
            title,
            source_url,
            source_file_name,
            source_file,
            sheet_name,
            column_name,
            column_wholesale,
            column_recommended,
            column_stock,
            wholesale_multiplier,
            recommended_multiplier,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            title,
            source_url,
            (file_name or "").strip(),
            sqlite3.Binary(file_bytes) if file_bytes else None,
            primary_sheet["sheet_name"],
            primary_sheet["column_name"],
            primary_sheet.get("column_wholesale"),
            primary_sheet.get("column_recommended"),
            primary_sheet.get("column_stock"),
            wholesale_multiplier,
            recommended_multiplier,
            created_at,
        ),
    )
    handler_id = cursor.lastrowid

    for position, config in enumerate(sheet_configs):
        cursor.execute(
            """
            INSERT INTO supplier_handler_sheets (
                handler_id,
                sheet_name,
                position,
                column_name,
                column_wholesale,
                column_recommended,
                column_stock,
                enabled
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 1)
            """,
            (
                handler_id,
                config["sheet_name"],
                position,
                config["column_name"],
                config.get("column_wholesale"),
                config.get("column_recommended"),
                config.get("column_stock"),
            ),
        )
    connection.commit()
    connection.close()

    handler = _get_supplier_handler(handler_id)
    try:
        rows = _refresh_handler_rows(handler)
        handler = _get_supplier_handler(handler_id) or handler
    except ValueError as error:
        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute("DELETE FROM supplier_handlers WHERE id = ?", (handler_id,))
        connection.commit()
        connection.close()
        return jsonify({"error": str(error)}), 400

    handler["rows"] = rows
    connection = get_db_connection()
    handler["changes"] = _get_recent_supplier_changes(connection, handler_id)
    connection.close()
    return jsonify({"handler": _serialize_handler(handler)})


@app.route("/api/suppliers/preview-source", methods=["POST"])
def preview_supplier_source():
    content_type = (request.content_type or "").lower()
    use_form = "multipart/form-data" in content_type or "application/x-www-form-urlencoded" in content_type

    handler = None
    file_bytes = None
    file_name = ""
    source_url = ""
    handler_id_value = None

    if use_form:
        form = request.form
        source_url = (form.get("source_url") or "").strip()
        handler_id_value = form.get("handler_id")
        file_storage = request.files.get("source_file")
        if file_storage and file_storage.filename:
            file_bytes = file_storage.read()
            file_name = file_storage.filename or ""
    else:
        data = request.get_json(force=True) or {}
        source_url = (data.get("source_url") or "").strip()
        handler_id_value = data.get("handler_id")

    if handler_id_value not in (None, ""):
        try:
            handler_id = int(handler_id_value)
        except (TypeError, ValueError):
            return jsonify({"error": "Некорректный идентификатор обработчика."}), 400
        handler = _get_supplier_handler(handler_id)
        if handler is None:
            return jsonify({"error": "Обработчик не найден."}), 404

    try:
        if file_bytes:
            workbook = _load_workbook_from_bytes(file_bytes, file_name=file_name)
        elif source_url:
            workbook = _download_supplier_workbook(source_url)
        elif handler is not None:
            workbook = _load_handler_workbook(handler)
        else:
            return jsonify({"error": "Укажите ссылку или загрузите файл."}), 400
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    response: Dict[str, Any] = {"sheets": _sheet_preview(workbook)}
    if handler is not None:
        sheets_payload: List[Dict[str, Any]] = []
        sheets_data = handler.get("sheets") or []
        if sheets_data:
            for sheet in sheets_data:
                sheets_payload.append(
                    {
                        "sheet_name": sheet.get("sheet_name"),
                        "mapping": {
                            "name": sheet.get("column_name"),
                            "wholesale_price": sheet.get("column_wholesale"),
                            "recommended_price": sheet.get("column_recommended"),
                            "stock": sheet.get("column_stock"),
                        },
                    }
                )
        else:
            sheets_payload.append(
                {
                    "sheet_name": handler.get("sheet_name"),
                    "mapping": {
                        "name": handler.get("column_name"),
                        "wholesale_price": handler.get("column_wholesale"),
                        "recommended_price": handler.get("column_recommended"),
                        "stock": handler.get("column_stock"),
                    },
                }
            )
        response["handler"] = {
            "id": handler["id"],
            "sheets": sheets_payload,
        }

    return jsonify(response)


def _update_supplier_handler(handler_id: int, handler: Dict[str, Any]):
    content_type = (request.content_type or "").lower()
    use_form = "multipart/form-data" in content_type or "application/x-www-form-urlencoded" in content_type

    if use_form:
        data_source = request.form
        file_storage = request.files.get("source_file")
        file_bytes = None
        file_name = ""
        if file_storage and file_storage.filename:
            file_bytes = file_storage.read()
            file_name = file_storage.filename or ""
    else:
        data_source = request.get_json(force=True) or {}
        file_bytes = None
        file_name = ""

    title = (data_source.get("title") or "").strip()
    source_url = (data_source.get("source_url") or "").strip()
    sheet_configs_raw = data_source.get("sheets")
    try:
        sheet_configs = _parse_sheet_configs(sheet_configs_raw, require_selection=True)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    try:
        wholesale_multiplier = _parse_multiplier(
            data_source.get("wholesale_multiplier"), default=handler.get("wholesale_multiplier") or 1.0
        )
    except ValueError as error:
        return jsonify({"error": f"Курс для опта: {error}"}), 400

    try:
        recommended_multiplier = _parse_multiplier(
            data_source.get("recommended_multiplier"), default=handler.get("recommended_multiplier") or 1.0
        )
    except ValueError as error:
        return jsonify({"error": f"Курс для РРЦ: {error}"}), 400

    if not title:
        return jsonify({"error": "Название обработчика обязательно."}), 400

    existing_file = handler.get("source_file")
    if existing_file is not None and not isinstance(existing_file, (bytes, bytearray)):
        existing_file = bytes(existing_file)

    final_file_bytes = file_bytes if file_bytes is not None else existing_file
    final_file_name = (file_name or handler.get("source_file_name") or "").strip()
    final_source_url = source_url

    if not final_source_url and not final_file_bytes:
        return jsonify({"error": "Укажите ссылку или загрузите файл."}), 400

    candidate_handler = dict(handler)
    candidate_handler.update(
        {
            "title": title,
            "source_url": final_source_url,
            "source_file_name": final_file_name,
            "source_file": final_file_bytes,
            "sheets": sheet_configs,
            "wholesale_multiplier": wholesale_multiplier,
            "recommended_multiplier": recommended_multiplier,
        }
    )

    try:
        if file_bytes:
            workbook = _load_workbook_from_bytes(file_bytes, file_name=file_name)
        elif final_source_url:
            workbook = _download_supplier_workbook(final_source_url)
        elif final_file_bytes:
            workbook = _load_workbook_from_bytes(final_file_bytes, file_name=final_file_name)
        else:
            return jsonify({"error": "Укажите ссылку или загрузите файл."}), 400
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    sheet_names = set(workbook.sheetnames)
    for config in sheet_configs:
        if config["sheet_name"] not in sheet_names:
            return jsonify({"error": f"Лист «{config['sheet_name']}» не найден в файле."}), 400

    primary_sheet = sheet_configs[0]
    candidate_handler.update(
        {
            "sheet_name": primary_sheet["sheet_name"],
            "column_name": primary_sheet["column_name"],
            "column_wholesale": primary_sheet.get("column_wholesale"),
            "column_recommended": primary_sheet.get("column_recommended"),
            "column_stock": primary_sheet.get("column_stock"),
        }
    )

    try:
        rows = _refresh_handler_rows(candidate_handler)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        "DELETE FROM supplier_handler_sheets WHERE handler_id = ?",
        (handler_id,),
    )
    for position, config in enumerate(sheet_configs):
        cursor.execute(
            """
            INSERT INTO supplier_handler_sheets (
                handler_id,
                sheet_name,
                position,
                column_name,
                column_wholesale,
                column_recommended,
                column_stock,
                enabled
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 1)
            """,
            (
                handler_id,
                config["sheet_name"],
                position,
                config["column_name"],
                config.get("column_wholesale"),
                config.get("column_recommended"),
                config.get("column_stock"),
            ),
        )
    cursor.execute(
        """
        UPDATE supplier_handlers
        SET
            title = ?,
            source_url = ?,
            source_file_name = ?,
            source_file = ?,
            sheet_name = ?,
            column_name = ?,
            column_wholesale = ?,
            column_recommended = ?,
            column_stock = ?,
            wholesale_multiplier = ?,
            recommended_multiplier = ?
        WHERE id = ?
        """,
        (
            title,
            final_source_url,
            final_file_name,
            sqlite3.Binary(final_file_bytes) if final_file_bytes else None,
            primary_sheet["sheet_name"],
            primary_sheet["column_name"],
            primary_sheet.get("column_wholesale"),
            primary_sheet.get("column_recommended"),
            primary_sheet.get("column_stock"),
            wholesale_multiplier,
            recommended_multiplier,
            handler_id,
        ),
    )
    connection.commit()
    handler_changes = _get_recent_supplier_changes(connection, handler_id)
    connection.close()

    updated_handler = _get_supplier_handler(handler_id) or candidate_handler
    updated_handler["rows"] = rows
    updated_handler["changes"] = handler_changes
    return jsonify({"handler": _serialize_handler(updated_handler)})


@app.route("/api/suppliers/<int:handler_id>", methods=["GET", "PUT"])
def supplier_handler_detail(handler_id: int):
    handler = _get_supplier_handler(handler_id)
    if handler is None:
        return jsonify({"error": "Обработчик не найден."}), 404

    if request.method == "PUT":
        return _update_supplier_handler(handler_id, handler)

    rows: Optional[List[Dict[str, Any]]] = None
    if _handler_has_mapping(handler):
        should_refresh = False
        refreshed_at_raw = handler.get("last_refreshed_at")
        if not refreshed_at_raw:
            should_refresh = True
        else:
            try:
                refreshed_at = datetime.fromisoformat(refreshed_at_raw)
                if refreshed_at <= datetime.utcnow() - timedelta(
                    minutes=SUPPLIER_AUTO_REFRESH_MINUTES
                ):
                    should_refresh = True
            except ValueError:
                should_refresh = True
        if should_refresh:
            try:
                rows = _refresh_handler_rows(handler)
                handler = _get_supplier_handler(handler_id) or handler
            except ValueError as error:
                handler["error"] = str(error)

    connection = get_db_connection()
    if rows is None:
        rows = _get_cached_supplier_rows(connection, handler_id)
    changes = _get_recent_supplier_changes(connection, handler_id)
    connection.close()
    handler["rows"] = rows
    handler["changes"] = changes
    return jsonify({"handler": _serialize_handler(handler)})


@app.route("/api/suppliers/<int:handler_id>/refresh", methods=["POST"])
def supplier_handler_refresh(handler_id: int):
    handler = _get_supplier_handler(handler_id)
    if handler is None:
        return jsonify({"error": "Обработчик не найден."}), 404

    try:
        rows = _refresh_handler_rows(handler)
        handler = _get_supplier_handler(handler_id) or handler
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    handler["rows"] = rows
    connection = get_db_connection()
    handler["changes"] = _get_recent_supplier_changes(connection, handler_id)
    connection.close()
    return jsonify({"handler": _serialize_handler(handler)})


@app.route("/api/suppliers/refresh-all", methods=["POST"])
def supplier_handlers_refresh_all():
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT id FROM supplier_handlers ORDER BY created_at DESC")
    handler_ids = [row["id"] for row in cursor.fetchall()]
    connection.close()

    if not handler_ids:
        return jsonify({"updated": 0, "failed": []})

    updated = 0
    failed = []

    for handler_id in handler_ids:
        handler = _get_supplier_handler(handler_id)
        if handler is None:
            failed.append({"id": handler_id, "error": "Обработчик не найден."})
            continue
        if not _handler_has_mapping(handler):
            failed.append({"id": handler_id, "error": "Не выбраны колонки."})
            continue
        try:
            _refresh_handler_rows(handler)
            updated += 1
        except ValueError as error:
            failed.append({"id": handler_id, "error": str(error)})

    return jsonify({"updated": updated, "failed": failed})


def _workbook_with_headers():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары"
    sheet.append(
        [
            "Категория",
            "Наименование",
            "Опт",
            "Цена продажи",
            "Остаток",
            "GTIN",
            "Дата поступления",
        ]
    )
    return workbook


@app.route("/export/template")
def export_template():
    workbook = _workbook_with_headers()
    sheet = workbook.active
    sheet.append(["Пример", "Товар", 0, None, 0, "", datetime.utcnow().date().isoformat()])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name="inventory_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/items")
def export_items():
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM items ORDER BY name")
    items = cursor.fetchall()
    connection.close()

    workbook = _workbook_with_headers()
    sheet = workbook.active
    for row in items:
        sheet.append(
            [
                row["category"],
                row["name"],
                row["wholesale_price"],
                row["sale_price"],
                row["quantity"],
                row["gtin"],
                row["date_received"],
            ]
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name="inventory_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _normalize_header_cell(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip().lower()
    return str(value).strip().lower()


def _row_is_empty(row) -> bool:
    if row is None:
        return True
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str):
            if cell.strip():
                return False
        elif cell != "":
            return False
    return True


def import_supplier_rows(rows):
    if not rows:
        raise ValueError("Файл не содержит данных.")

    total_rows = len(rows)
    start_index = 0
    while start_index < total_rows and _row_is_empty(rows[start_index]):
        start_index += 1

    if start_index >= total_rows:
        raise ValueError("Файл не содержит данных.")

    header_row_index = None
    for current_index in range(start_index, total_rows):
        row = rows[current_index] or ()
        normalized = [_normalize_header_cell(value) for value in row]
        if normalized[: len(IMPORT_EXPECTED_HEADERS)] == IMPORT_EXPECTED_HEADERS:
            header_row_index = current_index
            break

    if header_row_index is None:
        raise ValueError("Неверный шаблон файла.")

    if header_row_index + 1 >= total_rows:
        raise ValueError("Файл не содержит данных.")

    return rows[header_row_index + 1 :]


@app.route("/import/items", methods=["POST"])
def import_items():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Файл не найден."}), 400

    try:
        workbook = load_workbook(file, data_only=True)
    except Exception:
        return jsonify({"error": "Не удалось прочитать файл. Используйте шаблон."}), 400

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    try:
        rows = import_supplier_rows(rows)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    connection = get_db_connection()
    cursor = connection.cursor()
    imported = 0

    for row in rows:
        if row is None:
            continue
        category, name, wholesale_price, sale_price, quantity, gtin, date_received = (row + (None,) * 7)[:7]
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        try:
            wholesale_price_value = float(wholesale_price)
        except (TypeError, ValueError):
            continue
        try:
            quantity_value = int(quantity or 0)
        except (TypeError, ValueError):
            quantity_value = 0

        sale_price_value = None
        if sale_price not in (None, ""):
            try:
                sale_price_value = float(sale_price)
            except (TypeError, ValueError):
                sale_price_value = None

        date_value = datetime.utcnow().date().isoformat()
        if date_received not in (None, ""):
            try:
                date_value = datetime.fromisoformat(str(date_received)).date().isoformat()
            except ValueError:
                date_value = datetime.utcnow().date().isoformat()

        cursor.execute("SELECT id, quantity FROM items WHERE name = ?", (name,))
        existing = cursor.fetchone()
        if existing:
            cursor.execute(
                """
                UPDATE items
                SET category = ?, wholesale_price = ?, sale_price = ?, quantity = ?, gtin = ?, date_received = ?
                WHERE id = ?
                """,
                (
                    category,
                    wholesale_price_value,
                    sale_price_value,
                    quantity_value,
                    gtin,
                    date_value,
                    existing["id"],
                ),
            )
            delta = quantity_value - existing["quantity"]
            if delta:
                log_history(existing["id"], delta, "Импорт", connection=connection)
        else:
            cursor.execute(
                """
                INSERT INTO items (category, name, wholesale_price, sale_price, quantity, gtin, date_received)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    category,
                    name,
                    wholesale_price_value,
                    sale_price_value,
                    quantity_value,
                    gtin,
                    date_value,
                ),
            )
            item_id = cursor.lastrowid
            if quantity_value:
                log_history(item_id, quantity_value, "Импорт", connection=connection)
        imported += 1

    connection.commit()
    connection.close()
    return jsonify({"imported": imported})


@app.route("/health")
def health_check():
    return {"status": "ok"}


if __name__ == "__main__":
    init_db()
    app.run(debug=True)
else:
    init_db()
